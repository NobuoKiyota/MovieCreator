import openpyxl
import json
import os

# Paths are relative to this script's own location, not a hardcoded drive letter, so this runs
# unmodified on any PC regardless of where the repo is checked out (see CLAUDE.md multi-PC rules).
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(PROJECT_ROOT, "Excels", "ParameterRanges.xlsx")
out_json = os.path.join(PROJECT_ROOT, "data", "param_ranges.json")

GEN_SHEET_NAME = "Generator Params"
FX_SHEET_NAME = "Common FX Params"

if not os.path.exists(excel_file):
    print(f"Error: Excel file does not exist at {excel_file}!")
    print("Run `node scripts/seed_param_ranges.mjs` first to create it.")
    exit(1)

print("Loading Excel workbook...")
wb = openpyxl.load_workbook(excel_file, data_only=True)

for required in (GEN_SHEET_NAME, FX_SHEET_NAME):
    if required not in wb.sheetnames:
        print(f"Error: '{required}' sheet not found in the workbook!")
        exit(1)

# Only Min/Max (Actual) - columns E/F on the Generator Params sheet, D/E on Common FX Params - are
# read here. Min/Max (Reference) columns are documentation only (refreshed by
# scripts/seed_param_ranges.mjs from the current code, never read back into the app).

generator_params = {}
ws = wb[GEN_SHEET_NAME]
for row in ws.iter_rows(min_row=2, values_only=True):
    layer_type, param_name, label, step, min_actual, max_actual, min_ref, max_ref = row[:8]
    if not layer_type or not param_name:
        continue
    if min_actual is None or max_actual is None:
        continue
    generator_params.setdefault(str(layer_type).strip(), {})[str(param_name).strip()] = {
        "min": min_actual,
        "max": max_actual
    }

fx_params = {}
ws = wb[FX_SHEET_NAME]
for row in ws.iter_rows(min_row=2, values_only=True):
    param_name, label, step, min_actual, max_actual, min_ref, max_ref = row[:7]
    if not param_name:
        continue
    if min_actual is None or max_actual is None:
        continue
    fx_params[str(param_name).strip()] = {"min": min_actual, "max": max_actual}

mapping = {"generatorParams": generator_params, "fxParams": fx_params}

os.makedirs(os.path.dirname(out_json), exist_ok=True)
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(mapping, f, ensure_ascii=False, indent=2)

total_gen_entries = sum(len(v) for v in generator_params.values())
print("Parameter ranges exported successfully to:", out_json)
print(f"Generator types: {len(generator_params)}, total (type, param) entries: {total_gen_entries}")
print(f"Common FX param entries: {len(fx_params)}")
