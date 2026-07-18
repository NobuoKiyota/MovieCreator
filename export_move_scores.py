import openpyxl
import json
import os

# Paths are relative to this script's own location, not a hardcoded drive letter, so this runs
# unmodified on any PC regardless of where the repo is checked out (see CLAUDE.md multi-PC rules).
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(PROJECT_ROOT, "Excels", "PresetLayerOpinionSheet.xlsx")
out_json = os.path.join(PROJECT_ROOT, "data", "move_scores.json")

# Display name (Excel row 3 header, column order) -> internal layer type code
# (matches index.html's #layer-type-select option order 1:1)
LAYER_NAME_TO_TYPE = {
    "Sine Wave": "sine-wave",
    "Noise Wave": "noise-wave",
    "Firefly Particles": "particles",
    "Lissajous Geometry": "geometry",
    "Growing Sketch": "growing-sketch",
    "Neon Rain": "rain",
    "Meteor Shower": "meteor",
    "Pulse Ripples": "ripple",
    "Audio Spectrum": "spectrum",
    "3D Glowing Cube": "cube-3d",
    "Neon Lightning": "lightning",
    "Neon Fog": "fog",
    "Cyber Flame": "flame",
    "Neon Snowflake": "snowflake",
    "Neon Spirograph": "spirograph",
    "Aurora Curtain": "aurora",
    "Dry Ice Smoke": "dry-ice",
    "3D Shape Particles": "shape-3d-particles",
    "Lighthouse Beacon": "lighthouse",
    "Shockwave Burst": "shockwave-burst",
    "Glass Crack": "glass-crack",
}

if not os.path.exists(excel_file):
    print(f"Error: Excel file does not exist at {excel_file}!")
    exit(1)

print("Loading Excel workbook...")
wb = openpyxl.load_workbook(excel_file, data_only=True)
if "Preset Layers Opinion Sheet" not in wb.sheetnames:
    print("Error: 'Preset Layers Opinion Sheet' not found in the workbook!")
    exit(1)

ws = wb["Preset Layers Opinion Sheet"]

# Row 3: layer display names starting at column D (4), each spanning 3 columns (Score/Move/Comment)
layer_cols = {}  # display_name -> column index of its "Move" cell (Score col + 1)
col = 4
while col <= ws.max_column:
    name = ws.cell(row=3, column=col).value
    if name:
        layer_type = LAYER_NAME_TO_TYPE.get(str(name).strip())
        if layer_type:
            layer_cols[layer_type] = col + 1  # Move column is Score column + 1
        else:
            print(f"Warning: unrecognized layer display name '{name}' at column {col}, skipping.")
    col += 3

mapping = {t: {} for t in LAYER_NAME_TO_TYPE.values()}

for r in range(5, ws.max_row + 1):
    a_val = ws.cell(row=r, column=1).value
    if a_val and str(a_val).strip().startswith("---"):
        continue  # category banner row

    param_name = ws.cell(row=r, column=2).value
    if not param_name:
        continue
    param_name = str(param_name).strip()

    for layer_type, move_col in layer_cols.items():
        move_val = ws.cell(row=r, column=move_col).value
        if move_val is None or (isinstance(move_val, str) and move_val.strip() in ("", "-")):
            continue
        try:
            mapping[layer_type][param_name] = float(move_val)
        except (TypeError, ValueError):
            continue

os.makedirs(os.path.dirname(out_json), exist_ok=True)
with open(out_json, "w", encoding="utf-8") as f:
    json.dump(mapping, f, ensure_ascii=False, indent=2)

total_entries = sum(len(v) for v in mapping.values())
print("Move scores exported successfully to:", out_json)
print(f"Total layers: {len(mapping)}, total (layer, param) Move entries: {total_entries}")
