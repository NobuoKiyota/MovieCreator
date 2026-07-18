import openpyxl
import json
import os

# Paths are relative to this script's own location, not a hardcoded drive letter, so this runs
# unmodified on any PC regardless of where the repo is checked out (see CLAUDE.md multi-PC rules).
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(PROJECT_ROOT, "Excels", "PresetLayerOpinionSheet.xlsx")
out_json = os.path.join(PROJECT_ROOT, "data", "motion_mapping.json")

if not os.path.exists(excel_file):
    print(f"Error: Excel file does not exist at {excel_file}!")
    exit(1)

print("Loading Excel workbook...")
wb = openpyxl.load_workbook(excel_file, data_only=True)
if "Motion Mapping" not in wb.sheetnames:
    print("Error: 'Motion Mapping' sheet not found in the workbook!")
    exit(1)

ws = wb["Motion Mapping"]

mapping = {}
current_layer_id = None
motion_columns = ["SlowDriftUp", "SlowDriftDown", "PingPong", "PulseDecay", "SineLFO", "StepHold"]

# Parse rows
for r in range(5, ws.max_row + 1):
    val_a = ws.cell(row=r, column=1).value
    
    # Check if this row is a Layer Header Banner (starts with '■')
    if val_a and str(val_a).startswith("■"):
        banner_text = str(val_a)
        if "(" in banner_text and ")" in banner_text:
            current_layer_id = banner_text.split("(")[-1].replace(")", "").strip()
            mapping[current_layer_id] = {}
        continue
        
    if not val_a or val_a == "Category":
        continue
        
    if current_layer_id:
        param_name = ws.cell(row=r, column=2).value
        
        if param_name:
            param_name = str(param_name).strip()
            allowed_motions = []
            
            for idx, col in enumerate(range(4, 10)):
                cell_val = ws.cell(row=r, column=col).value
                if cell_val:
                    cell_val_str = str(cell_val).strip().lower()
                    if cell_val_str in ['o', 'y', '1', 'yes', 'true', 'ok', 'a', 'b']:
                        allowed_motions.append(motion_columns[idx])
            
            if allowed_motions:
                mapping[current_layer_id][param_name] = allowed_motions

# Ensure data directory exists
os.makedirs(os.path.dirname(out_json), exist_ok=True)

with open(out_json, "w", encoding="utf-8") as f:
    json.dump(mapping, f, ensure_ascii=False, indent=2)

print("Motion Mapping exported successfully to:", out_json)
print(f"Total layers mapped: {len(mapping)}")
